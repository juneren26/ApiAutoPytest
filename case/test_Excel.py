#!/usr/bin/env python
# -*- coding: utf-8 -*-
import allure
import openpyxl
import pytest
from VAR.VAR import *
from api_keyword.api_key import ApiKey
from data_driver.excel_read import read_excel


# setup_module在当前文件中，在所有测试用例执行之前与之后执行
def setup_module():
    # 1.定义全局变量，当前文件的函数都可以访问
    global ak, excel, sheet, excel_path, all_val
    # 2.实例化工具类
    ak = ApiKey()
    # 3.初始化excel文件
    excel_path = EXCEL_PATH
    excel = openpyxl.load_workbook(excel_path)
    sheet = excel['Sheet1']
    # 4. 参数化变量存储字典，临时数据库
    all_val = {}


@pytest.mark.parametrize('data', read_excel())
def test01(data):
    # # 动态生成标题
    # allure.dynamic.title(data[11])
    # 如果存在自定义标题
    if data[11] is not None:
        # 动态生成标题
        allure.dynamic.title(data[11])

    if data[16] is not None:
        # 动态获取story模块名
        allure.dynamic.story(data[16])

    if data[17] is not None:
        # 动态获取feature模块名
        allure.dynamic.feature(data[17])

    if data[18] is not None:
        # 动态获取备注信息
        allure.dynamic.description(data[18])

    if data[19] is not None:
        # 动态获取级别信息(blocker、critical、normal、minor、trivial)
        allure.dynamic.severity(data[19])

    # 行数
    r = data[0] + 1

    # ==============Excel数据解析==============
    try:
        dict_data = {
            'url': data[1] + data[2],
            'params': eval(data[4]),
            'headers': eval(data[5]),
            data[7]: eval(data[6])
        }

    except:
        print("=============实际结果=================")
        print("接口请求格式有误，请检查url、params、headers、data、参数类型")
        sheet.cell(r, 11).value = "请求参数有误，请检查"
        raise

    # 发起请求
    # 反射+解包，等同于 requests.get()
    # 这里不会有异常，因此用try无效，只能根据报告里的参数去检查
    # try:
    #     res = getattr(ak,data[3])(**dict_data)
    # except:
    #     print("=============实际结果=================")
    #     print("请求参数有误，请检查data中传递的参数是否符合接口要求")
    #     sheet.cell(r, 11).value = "请求参数有误，请检查"
    #     raise

    res = getattr(ak, data[3])(**dict_data)
    # c = getattr(a,'set')(a='1',b='2')
    # 打印响应报文
    print(res.text)

    # =================Json提取器=================
    if data[12] is not None:
        # 遍历分割JSON提取_引用名称
        varStr = data[12]
        # 用分号分割varStr字符串，并保存到列表
        varStrList = varStr.split(';')
        # 获取列表长度
        length = len(varStrList)
        # print(length)

        # 遍历分割JSON表达式
        jsonStr = data[13]
        jsonList = jsonStr.split(';')

        # 循环输出列表值
        for i in range(length):
            # 获取JSON提取_引用名称
            key = varStrList[i]
            # print("获取key值")
            # print(key)

            # json表达式获取
            jsonExp = jsonList[i]
            # print("获取表达式")
            # print(jsonExp)

            # 字典值获取
            valueJson = ak.get_text(res.text, jsonExp)
            # print(valueJson)

            # 持续添加参数，只要参数名不重复，重复的后面就会覆盖前面的参数
            all_val[key] = valueJson

    # 报文结果校验
    try:
        # 实际结果
        result = None
        result = ak.get_text(res.text, data[8])
        if result == data[9]:
            sheet.cell(r, 11).value = "通过"
        else:
            sheet.cell(r, 11).value = "不通过"
        excel.save(excel_path)
    except:
        print("=============实际结果=================")
        print("jsonpath表达式有误，请检查")
        sheet.cell(r, 11).value = "jsonpath表达式有误，请检查"
        excel.save(excel_path)
    finally:
        assert result == data[9]

    # 数据库结果校验
    # 如果存在数据库检查并且需要动态关联接口参数
    if data[21] is not None:
        try:
            # 拼接sql
            str1 = data[20]
            # 遍历分割数据库变量
            sqlStr = data[21]
            sqlList = sqlStr.split(';')  # ["all_val['VAR_UID']", "all_val['VAR_UNAME']"]
            list1 = []
            # 获取列表长度
            length = len(sqlList)
            # 循环解析all_val[]值并填入list1
            for i in range(length):
                list1.append(eval(sqlList[i]))
            sql = str1.format(*list1)

            # 实际结果
            sql_check = None
            sql_check = ak.sqlCheck(sql)
            if (sql_check == data[22]):
                sheet.cell(r, 11).value = "通过"
            else:
                sheet.cell(r, 11).value = "不通过"
            excel.save(excel_path)
        except:
            print("=============实际结果=================")
            print("sql有误，或者查询结果为空，请检查")
            sheet.cell(r, 11).value = "sql有误，请检查"
            excel.save(excel_path)
        finally:
            assert sql_check == data[22]
    # 数据库结果校验
    # 如果存在数据库检查，不需要动态关联接口参数
    elif data[20] is not None:
        try:
            # 实际结果
            sql_check = None
            sql = data[20]
            sql_check = ak.sqlCheck(sql, 1)
            if (sql_check == data[22]):
                sheet.cell(r, 11).value = "通过"
            else:
                sheet.cell(r, 11).value = "不通过"
            excel.save(excel_path)
        except:
            print("=============实际结果=================")
            print("sql有误，请检查")
            sheet.cell(r, 11).value = "sql有误，请检查"
            excel.save(excel_path)
        finally:
            assert sql_check == data[22]

    print("=============参数信息=================")
    print(all_val)
